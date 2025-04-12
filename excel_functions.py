from openpyxl import load_workbook
class ExcelReader:

    def __init__(self,excel_file,sheet_name):
        self.file = excel_file
        self.sheet = sheet_name

    """ Fetch the row count from excel file"""
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    """ Fetch the column count from excel file"""
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    """Read the data from excel file"""
    def read_data(self,row_number,column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        data = sheet.cell(row=row_number, column=column_number).value
        return data

    """Write the data from excel file"""
    def write_data(self,row_number,column_number,data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number,column=column_number).value = data
        workbook.save(self.file)
